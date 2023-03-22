import re

import openpyxl

wrkbk = openpyxl.load_workbook("input.xlsx")
sh = wrkbk.active
print("Start")
# for i in range(1, sh.max_row + 1):
pgm_titles = []
output_rows = []
program_name = ''

# create a new Excel workbook
workbook = openpyxl.Workbook()

# select the active worksheet
worksheet = workbook.active

headings = ["County", "Program Name", "Family Size", "Income limit %", "Income limit", "Year"]
worksheet.append(headings)

county = ''
for row in sh.iter_rows(min_row=2, max_row=sh.max_row):
    if not row[0]:
        continue
    first_col = str(row[0].value)
    row_1 = row[1]

    if first_col and row[1].value is None:
        program_name = first_col
        county = program_name.split('-')[0]
        print(f"Processing {county}")
        continue  # Skip to the next row

    if first_col and first_col.strip() == 'Persons in Family':
        for j in range(1, len(row)):
            j_val = row[j].value
            print(f"Processing {j_val}")
            pgm_titles.append(int(re.search("\((\d+)%\)", j_val).group(1)))
    else:
        # Otherwise, assume it's a program detail row and extract the relevant information
        family_size = row[0].value
        # Iterate over the columns in the program detail row to extract the income limit type and value
        for j in range(1, len(row)):
            income_limit_value = row[j].value
            # Create a new output row with the extracted information
            output_row = [county, program_name, family_size, income_limit_value, pgm_titles[j-1], 2023]
            worksheet.append(output_row)

workbook.save('output.xlsx')
